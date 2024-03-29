from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime, timedelta
import sys

'''表格格式不变时通常不需要换,代表有效数据起始的行和列'''
ROW_START = 5
COL_START = ord('V')-ord('A')+1
DAY_START_TIME = datetime.strptime("10:00", "%H:%M")
DAY_END_TIME = datetime.strptime("19:00", "%H:%M")
LAUNCH_BONUS_END_TIME = datetime.strptime("10:30", "%H:%M")


def judge(cell):
    font_late = Font(color="FF0000")
    font_leave_early = Font(color="FF0000")
    # font_bonus = Font(color="912CEE")
    font_holiday = Font(color="00FF00")

    leave_early_delta = timedelta(hours=9)
    bonus_delta = timedelta(hours=11)
    should_give_bonus = False
    should_give_launch_bonus = False
    leave_eary_judge = False
    late = False
    s = cell.value
    v = s.split(r";")

    if len(v) == 3:
        # 存在正常的上下班打卡记录
        s = v[0].strip()
        start = datetime.strptime(s, "%H:%M")
        if start > DAY_START_TIME:
            cell.font = font_late
            late = True
            if start <= LAUNCH_BONUS_END_TIME:
                start = DAY_START_TIME  # 迟到后上班时间按10:00计算
        e = v[1].strip()
        if "次日" in e:
            # 异常数据处理
            e = "23:59"
        end = datetime.strptime(e, "%H:%M")
        if end-start < leave_early_delta:
            cell.font = font_leave_early
            leave_eary_judge = True
        if end-start >= bonus_delta:
            should_give_bonus = True
            # cell.font = font_bonus

        if start <= LAUNCH_BONUS_END_TIME and not leave_eary_judge:
            should_give_launch_bonus = True
        cell.value = "{}\n{}".format(s, e)
    else:
        if "休息" in s:
            cell.font = font_holiday
        else:
            cell.font = font_late
            late = True
            cell.value = "{}\n".format(v[0].strip())

    return should_give_bonus, should_give_launch_bonus, late, leave_eary_judge


def find_holiday(ws, row_max, col_max):
    holidays_idx = []
    for c in range(COL_START, col_max+1):
        holiday_cnt = len(
            [r for r in range(ROW_START, row_max+1) if "休息" in ws.cell(r, c).value])
        if holiday_cnt >= 1:
            holidays_idx.append(c)
    print("本月休息日为:")
    month_days = col_max-COL_START
    wokr_days = month_days-len(holidays_idx)+1
    print([x-COL_START+1 for x in holidays_idx])
    print(wokr_days)
    return holidays_idx


def handle(in_file, out_file):
    wb = load_workbook(in_file)
    ws = wb[wb.sheetnames[0]]
    row_max, col_max = ws.max_row, ws.max_column
    print(ws.cell(ROW_START, COL_START).value)

    holiday_idx = find_holiday(ws, row_max, col_max)
    ws.cell(ROW_START-1, col_max+1).value = "晚餐补助"
    ws.cell(ROW_START-1, col_max+2).value = "午餐补助"
    ws.cell(ROW_START-1, col_max+3).value = "迟到"
    ws.cell(ROW_START-1, col_max+4).value = "早退"
    for r in range(ROW_START, row_max+1):
        print(ws.cell(r, 1).value)
        bonus_cnt = 0
        launch_bonus_cnt = 0
        early_quit = 0
        late = 0
        for c in range(COL_START, col_max+1):
            cell = ws.cell(r, c)
            should_give_bonus, should_give_launch_bonus, late_judge, leave_early_quit_judge = judge(
                cell)
            if c not in holiday_idx:
                bonus_cnt += 1 if should_give_bonus else 0
                launch_bonus_cnt += 1 if should_give_launch_bonus else 0
                early_quit += 1 if leave_early_quit_judge else 0
                late += 1 if late_judge else 0
        ws.cell(r, col_max+1).value = bonus_cnt
        ws.cell(r, col_max+2).value = launch_bonus_cnt
        ws.cell(r, col_max+3).value = late
        ws.cell(r, col_max+4).value = early_quit

    wb.save(out_file)


if __name__ == '__main__':
    if len(sys.argv) == 3:
        in_file = sys.argv[1]
        out_file = sys.argv[2]
    else:
        in_file = "D:/pypj/staffattendance/a.xlsx"
        out_file = "D:/pypj/staffattendance/b.xlsx"
    print(in_file, out_file)
    handle(in_file, out_file)
