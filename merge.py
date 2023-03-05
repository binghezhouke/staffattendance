from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime, timedelta
import sys


def handle(in_file, out_file):
    # From G -> W
    info_map = ['E', 'G', 'F', 'K', 'Z', 'J', 'Z', 'Z',
                'Z', 'N', 'P', 'O', 'L', 'M', 'Q', 'S', 'T']
    wb_in = load_workbook(in_file)
    ws_in = wb_in[wb_in.sheetnames[0]]
    row_max, col_max = ws_in.max_row, ws_in.max_column
    ws_in_start_row = 5

    def find_row_idx_by_name(name, ws, s, e):
        for r in range(s, e+1):
            if name == ws.cell(r, 1).value.strip():
                return r

    wb_out = load_workbook(out_file)
    ws_out = wb_out[wb_out.sheetnames[0]]
    o_row_max, o_col_max = ws_out.max_row, ws_out.max_column
    ws_out_start_row = 3
    ws_out_start_col = ord('G')-ord('A')+1
    for i in range(ws_out_start_row, o_row_max+1):
        name = ws_out.cell(i, 2).value.strip()
        print(name)
        row_idx_in_ori = find_row_idx_by_name(
            name, ws_in, ws_in_start_row, row_max)
        if not row_idx_in_ori:
            print("{} not found".format(name))
            continue
        for j in range(ws_out_start_col, ws_out_start_col+len(info_map)):
            col_idx_in = info_map[j-ws_out_start_col]
            if 'Z' != col_idx_in:
                in_value = ws_in.cell(
                    row_idx_in_ori, ord(col_idx_in)-ord('A')+1).value
                try:
                    in_value = float(in_value)
                except:
                    in_value = 0
                ws_out.cell(i, j).value = in_value
    wb_out.save(out_file+".out.xlsx")


if __name__ == '__main__':
    if len(sys.argv) == 3:
        in_file = sys.argv[1]
        out_file = sys.argv[2]
    else:
        in_file = "D:/pypj/staffattendance/a.xlsx"
        out_file = "D:/pypj/staffattendance/b.xlsx"
    print(in_file, out_file)
    handle(in_file, out_file)
